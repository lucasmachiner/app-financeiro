"""#!/usr/bin/env python3
# generate_xlsx.py
# Reads spreadsheets/expenses_template.csv and generates planilha_financeiro.xlsx with two sheets:
# - Expenses (data from CSV)
# - Resumo Mensal (12 months from 2025-11-01 to 2026-10-01) with formulas that distribute parcel payments

import csv
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

CSV_PATH = 'spreadsheets/expenses_template.csv'
OUT_XLSX = 'planilha_financeiro.xlsx'

# Read CSV
rows = []
with open(CSV_PATH, newline='', encoding='utf-8') as f:
    reader = csv.reader(f)
    header = next(reader)
    for r in reader:
        rows.append(r)

wb = Workbook()
# Expenses sheet
ws = wb.active
ws.title = 'Expenses'
ws.append(header)
for r in rows:
    ws.append(r)
# Format header bold (optional)

# Add formula column for Valor_Parcela if Parcelas provided (column F -> index 6)
# We'll assume columns: A Date, B Tipo, C Categoria, D Descrição, E Valor_Total, F Parcelas, G Valor_Parcela, H Status, I Observações
for row_idx in range(2, 2 + len(rows)):
    parcel_cell = f'F{{row_idx}}'
    valor_total_cell = f'E{{row_idx}}'
    valor_parc_cell = f'G{{row_idx}}'
    ws[valor_parc_cell] = f'=IF({parcel_cell}>0, {valor_total_cell}/{parcel_cell}, "")'

# Resumo Mensal sheet
rs = wb.create_sheet('Resumo Mensal')
start_year = 2025
start_month = 11
months = []
for i in range(12):
    y = (start_month - 1 + i) // 12 + start_year
    m = (start_month - 1 + i) % 12 + 1
    months.append((y, m))

# Header row: first cell blank then months as dates (1st of month)
rs.append([''] + [date(y, m, 1).isoformat() for (y, m) in months])
# Rows: Despesas, Receitas, Saldo Mensal, Saldo Acumulado
rs.append(['Despesas'] + ['' for _ in months])
rs.append(['Receitas'] + ['' for _ in months])
rs.append(['Saldo Mensal'] + ['' for _ in months])
rs.append(['Saldo Acumulado'] + ['' for _ in months])

# Insert formulas for Despesas/Receitas
# We'll use ranges up to row 1000; adjust if needed
for col_idx in range(2, 2 + len(months)):
    col_letter = get_column_letter(col_idx)
    header_cell = f'{col_letter}1'
    # Despesas formula
    desp_formula = (
        "=SUMPRODUCT(((YEAR(" + header_cell + " )-YEAR(Expenses!$A$2:$A$1000))*12 + MONTH(" + header_cell + " )-MONTH(Expenses!$A$2:$A$1000) + 1 >= 1) * "
        "(((YEAR(" + header_cell + " )-YEAR(Expenses!$A$2:$A$1000))*12 + MONTH(" + header_cell + " )-MONTH(Expenses!$A$2:$A$1000) + 1) <= Expenses!$F$2:$F$1000) * "
        "(Expenses!$B$2:$B$1000=\"Despesa\") * (Expenses!$E$2:$E$1000 / Expenses!$F$2:$F$1000))"
    )
    rs[f'{col_letter}2'] = desp_formula
    # Receitas formula (row 3)
    rec_formula = desp_formula.replace('Despesa', 'Receita')
    rs[f'{col_letter}3'] = rec_formula
    # Saldo Mensal = Receitas - Despesas (row 4)
    rs[f'{col_letter}4'] = f'={col_letter}3 - {col_letter}2'

# Saldo Acumulado (row 5): first month = Saldo Mensal, next months = previous acumulado + current saldo
# Row numbers: 2 Despesas,3 Receitas,4 Saldo Mensal,5 Saldo Acumulado
# First month col B
rs['B5'] = '=B4'
for col_idx in range(3, 2 + len(months)):
    col_letter = get_column_letter(col_idx)
    left_col = get_column_letter(col_idx - 1)
    rs[f'{col_letter}5'] = f'={left_col}5 + {col_letter}4'

# Save
wb.save(OUT_XLSX)
print('Generated', OUT_XLSX)
"""
